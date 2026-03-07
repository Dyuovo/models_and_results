import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference


def load_csv(path):
    df = pd.read_csv(path)
    df["times"] = pd.to_datetime(df["times"], errors="coerce", format="mixed")
    df = df.dropna(subset=["times"])
    return df


def write_data_sheet(wb, name, df):
    ws = wb.create_sheet(name)
    ws.append(["times", "time_of_day", "day_ahead_price", "real_time_price"])
    for row in df[["times", "day_ahead_price", "real_time_price"]].itertuples(index=False):
        ws.append([row[0], None, row[1], row[2]])
    max_row = ws.max_row
    for r in range(2, max_row + 1):
        ws[f"B{r}"] = f"=MOD(A{r},1)"
    ws["A1"].number_format = "yyyy-mm-dd hh:mm"
    ws["B1"].number_format = "h:mm"
    for r in range(2, max_row + 1):
        ws[f"A{r}"].number_format = "yyyy-mm-dd hh:mm"
        ws[f"B{r}"].number_format = "h:mm"
    return ws, max_row


def add_summary_sheet(wb, data_2025_rows, data_2026_rows):
    ws = wb.create_sheet("summary")
    ws.append(["time_of_day", "day_ahead_2025", "day_ahead_2026", "real_time_2025", "real_time_2026"])

    slots = pd.date_range("2000-01-01 00:00", "2000-01-01 23:45", freq="15min").time
    for i, t in enumerate(slots, start=2):
        ws[f"A{i}"] = t
        ws[f"A{i}"].number_format = "h:mm"
        ws[f"B{i}"] = f"=AVERAGEIF(data_2025!$B$2:$B${data_2025_rows},$A{i},data_2025!$C$2:$C${data_2025_rows})"
        ws[f"C{i}"] = f"=AVERAGEIF(data_2026!$B$2:$B${data_2026_rows},$A{i},data_2026!$C$2:$C${data_2026_rows})"
        ws[f"D{i}"] = f"=AVERAGEIF(data_2025!$B$2:$B${data_2025_rows},$A{i},data_2025!$D$2:$D${data_2025_rows})"
        ws[f"E{i}"] = f"=AVERAGEIF(data_2026!$B$2:$B${data_2026_rows},$A{i},data_2026!$D$2:$D${data_2026_rows})"

    chart1 = LineChart()
    chart1.title = "Average Day-Ahead Price by Time of Day"
    chart1.y_axis.title = "Price"
    chart1.x_axis.title = "Time of Day"
    data1 = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=97)
    cats = Reference(ws, min_col=1, min_row=2, max_row=97)
    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats)
    ws.add_chart(chart1, "G2")

    chart2 = LineChart()
    chart2.title = "Average Real-Time Price by Time of Day"
    chart2.y_axis.title = "Price"
    chart2.x_axis.title = "Time of Day"
    data2 = Reference(ws, min_col=4, max_col=5, min_row=1, max_row=97)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats)
    ws.add_chart(chart2, "G20")

    return ws


def main():
    path_2025 = r"c:\Users\Joey\Desktop\xmo\guangdong_2025_complete.csv"
    path_2026 = r"c:\Users\Joey\Desktop\xmo\guangdong_2026_complete.csv"
    out_xlsx = r"c:\Users\Joey\Desktop\xmo\avg_intraday_prices.xlsx"

    df_2025 = load_csv(path_2025)
    df_2026 = load_csv(path_2026)

    wb = Workbook()
    wb.remove(wb.active)

    _, rows_2025 = write_data_sheet(wb, "data_2025", df_2025)
    _, rows_2026 = write_data_sheet(wb, "data_2026", df_2026)
    add_summary_sheet(wb, rows_2025, rows_2026)

    wb.save(out_xlsx)


if __name__ == "__main__":
    main()
